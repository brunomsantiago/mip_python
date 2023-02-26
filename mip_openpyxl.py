import openpyxl

# Loading an excel file with sensitivity label (MIP Label)
file_with_mip_label = 'file_with_mip_label.xlsx'
workbook_with_mip_label = openpyxl.load_workbook(file_with_mip_label)

# Creating a new workbook and writing something in ht
new_workbook = openpyxl.Workbook()
worksheet = new_workbook.active
cell = worksheet.cell(row=1, column=1)
cell.value = 'Hello World'

# Copying custom properties from one workbook to another
for prop in workbook_with_mip_label.custom_doc_props.props:
    print(f"{prop.name}: {prop.value}")
    new_workbook.custom_doc_props.append(prop)

# Saving the new workbook
new_workbook.save('new_file_with_same_mip_label.xlsx')